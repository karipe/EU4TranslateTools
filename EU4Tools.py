# -*- coding: utf-8 -*-

import os
import os.path
import subprocess
import sys
import re
import unicodedata

cp_map = {
    0x80: 0x20AC,
    0x82: 0x201A,
    0x83: 0x0192,
    0x84: 0x201E,
    0x85: 0x2026,
    0x86: 0x2020,
    0x87: 0x2021,
    0x88: 0x02C6,
    0x89: 0x2030,
    0x8A: 0x0160,
    0x8B: 0x2039,
    0x8C: 0x0152,
    0x8E: 0x017D,
    0x91: 0x2018,
    0x92: 0x2019,
    0x93: 0x201C,
    0x94: 0x201D,
    0x95: 0x2022,
    0x96: 0x2013,
    0x97: 0x2014,
    0x98: 0x02DC,
    0x99: 0x2122,
    0x9A: 0x0161,
    0x9B: 0x203A,
    0x9C: 0x0153,
    0x9E: 0x017E,
    0x9F: 0x0178
}


def cp1252_to_ucs2(cp):
    if cp < 0:
        cp += 256
    if cp in cp_map:
        result = cp_map[cp].to_bytes(2, byteorder='little')
    else:
        result = cp.to_bytes(2, byteorder='little')
    return result


errornous_characters = set()


def convert_wide_text_to_escaped_wide_text(wide_text):
    result = bytearray()
    special_chars = [0xA4, 0xA3, 0xA7, 0x24, 0x5B, 0x00, 0x5C, 0x20, 0x0D, 0x0A,
                     0x22, 0x7B, 0x7D, 0x40, 0x80, 0x7E, 0xBD]
    for i in range(0, len(wide_text), 2):
        escape_char = 0x10
        high = wide_text[i + 1]
        low = wide_text[i]
        if high == 0:
            result += bytes([low, 0x00])
            continue

        if high in special_chars:
            escape_char += 2
            high -= 9
        if low in special_chars:
            escape_char += 1
            low += 15
        if high == 0x40 or low == 0x40:
            errornous_characters.add(wide_text[i:i + 2].decode('utf-16-le'))

        result.extend(escape_char.to_bytes(2, byteorder='little') + cp1252_to_ucs2(low) + cp1252_to_ucs2(high))

    return result

def convert_wide_text_to_escaped_text(wide_text):
    result = bytearray()
    special_chars = [0xA4, 0xA3, 0xA7, 0x24, 0x5B, 0x00, 0x5C, 0x20, 0x0D, 0x0A,
                     0x22, 0x7B, 0x7D, 0x40, 0x80, 0x7E, 0xBD]
    for i in range(0, len(wide_text), 2):
        escape_char = 0x10
        high = wide_text[i + 1]
        low = wide_text[i]
        if high == 0:
            result += bytes([low])
            continue

        if high in special_chars:
            escape_char += 2
            high -= 9
        if low in special_chars:
            escape_char += 1
            low += 15

        result.extend(escape_char.to_bytes(1, byteorder='little') + low.to_bytes(1, byteorder='little') + high.to_bytes(1, byteorder='little'))

    return result


def test(src):
    return convert_wide_text_to_escaped_wide_text(src.encode('utf-16-le'))


characters = set()
worldmap_characters = set()

FULL2HALF = dict((i + 0xFEE0, i) for i in range(0x21, 0x7F))
FULL2HALF[0x3000] = 0x20

def extract_characters(name, text):
    text = str(text).translate(FULL2HALF)
    for c in text:
        if c == '\n':
            continue
        characters.add(c)
        if name in [None, 'countries_l_english', 'prov_names_l_english', 'area_regions_l_english',
                    'cultures_phase4_l_english', 'prov_names_adj_l_english', 'regions_phase4_l_english',
                    'tags_phase4_l_english']:
            worldmap_characters.add(c)


def process_file(name, is_unified=False):
    print('Processing ' + name + '...')
    f = open('./original/yml/' + name + '.yml', "r", encoding='utf_8_sig')
    text = f.read()
    f.close()
    extract_characters(name, text)

    wide_text = text.encode('utf-16-le')
    escaped_wide_text = convert_wide_text_to_escaped_wide_text(wide_text)
    text = escaped_wide_text.decode('utf-16-le')
    os.makedirs('result/original_yml', exist_ok=True)
    f = open('./result/original_yml/' + name + '.yml', 'w', encoding='utf_8_sig')
    f.write(text)
    f.close()

    if is_unified:
        if os.path.isfile('./unified/yml/' + name + '.yml'):
            f = open('./unified/yml/' + name + '.yml', "r", encoding='utf_8_sig')
            text = f.read()
            f.close()
            extract_characters(name, text)

            wide_text = text.encode('utf-16-le')
            escaped_wide_text = convert_wide_text_to_escaped_wide_text(wide_text)
            text = escaped_wide_text.decode('utf-16-le')
            os.makedirs('result/unified_yml', exist_ok=True)
            f = open('./result/unified_yml/' + name + '.yml', 'w', encoding='utf_8_sig')
            f.write(text)
            f.close()
        else:
            os.makedirs('result/unified_yml', exist_ok=True)
            f = open('./result/unified_yml/' + name + '.yml', 'w', encoding='utf_8_sig')
            f.write(text)
            f.close()

def escape_yml(is_unified):
    if is_unified:
        files = list(set(os.listdir("./unified/yml") + os.listdir("./original/yml")))
    else:
        files = os.listdir("./original/yml")
    for file in files:
        file_name, file_ext = os.path.splitext(file)

        if file_ext == ".yml":
            process_file(file_name, is_unified)

    # Add default texts
    f = open('./default_source.txt', 'r', encoding='utf_8_sig')
    text = f.read()
    extract_characters(None, text)
    f.close()

    f = open('./ingame_source.txt', 'w', encoding='utf_8_sig')
    f.write(''.join(sorted(list(characters))))
    f.close()
    f = open('./worldmap_source.txt', 'w', encoding='utf_8_sig')
    f.write(''.join(sorted(list(worldmap_characters))))
    f.close()

def escape_txt(is_unified):
    if is_unified:
        processed_files = set()
        for category in os.listdir('./unified/txt'):
            if os.path.isdir('./unified/txt/%s' % category):
                for filename in os.listdir('./unified/txt/%s' % category):
                    processed_files.add((category, filename))
                    f = open('./unified/txt/%s/%s' % (category, filename), 'r', encoding='utf-8-sig')
                    text = f.read()
                    extract_characters('txt', text)
                    f.close()
                    wide_text = text.encode('utf-16-le')
                    text = convert_wide_text_to_escaped_text(wide_text)

                    os.makedirs('result/txt/%s' % (category), exist_ok=True)
                    f = open('./result/txt/%s/%s' % (category, filename), 'wb')
                    f.write(text)
                    f.close()

        for category in os.listdir('./original/txt'):
            if os.path.isdir('./original/txt/%s' % category):
                for filename in os.listdir('./original/txt/%s' % category):
                    if (category, filename) in processed_files:
                        continue

                    try:
                        f = open('./original/txt/%s/%s' % (category, filename), 'r', encoding='utf-8-sig')
                        text = f.read()
                        print(category, filename, 'utf-8-sig')
                    except UnicodeDecodeError as e:
                        f = open('./original/txt/%s/%s' % (category, filename), 'r', encoding='ISO-8859-1')
                        text = f.read()
                        print(category, filename, 'iso-8859-1')
                    extract_characters('txt', text)
                    f.close()
                    wide_text = text.encode('utf-16-le')
                    escaped_wide_text = convert_wide_text_to_escaped_wide_text(wide_text)
                    text = escaped_wide_text.decode('utf-16-le')

                    os.makedirs('result/txt/%s' % (category), exist_ok=True)
                    f = open('./result/txt/%s/%s' % (category, filename), 'w', encoding='utf-8-sig')
                    f.write(text)
                    f.close()
    else:
        for category in os.listdir('./original/txt'):
            if os.path.isdir('./original/txt/%s' % category):
                for filename in os.listdir('./original/txt/%s' % category):
                    try:
                        f = open('./original/txt/%s/%s' % (category, filename), 'r', encoding='utf-8-sig')
                        text = f.read()
                    except UnicodeDecodeError as e:
                        f = open('./original/txt/%s/%s' % (category, filename), 'r', encoding='ISO-8859-1')
                        text = f.read()
                    extract_characters('txt', text)
                    f.close()
                    wide_text = text.encode('utf-16-le')
                    escaped_wide_text = convert_wide_text_to_escaped_wide_text(wide_text)
                    text = escaped_wide_text.decode('utf-16-le')

                    os.makedirs('result/txt/%s' % (category), exist_ok=True)
                    f = open('./result/txt/%s/%s' % (category, filename), 'w', encoding='utf-8-sig')
                    f.write(text)
                    f.close()


def generate_bmfont(name, ext, source_file):
    os.makedirs('result/fonts', exist_ok=True)
    cmd_list = [
        'bmfont64.exe',
        '-c',"./original/bmfc/" + name + ext,
        '-o',"./result/fonts/" + name + ".fnt",
        '-t',source_file
    ]

    subprocess.call(' '.join(cmd_list))
    f = open('./result/fonts/' + name + '.fnt', 'r', encoding='utf_8')
    lines = f.readlines()
    f.close()
    f = open('./result/fonts/' + name + '.fnt', 'w', encoding='utf_8')

    if ext == '.bmfc':
        umlauts = {
            # 오류 우회
            '선': 'ퟴ',
            '휠': 'ퟵ',
            '젠': 'ퟶ',
            '츠': 'ퟷ',
            '유': 'ퟸ',
            '퀠': 'ퟹ',
            '술': 'ퟺ',
            '갠': 'ퟻ',
            ' ': '　',

            '년': [15],
            '일': [14],

            'A': 'ÀÁÂÃÄÅ',
            'C': 'Ç',
            'E': 'ÈÉÊË',
            'I': 'ÌÍÎÏ',
            'N': 'Ñ',
            'O': 'ÒÓÔÕÖ',
            'U': 'ÙÚÛÜ',
            'Y': 'ÝŸ',
            'S': 'Š',
            'Z': 'Ž',
            'a': 'àáâãäå',
            'c': 'ç',
            'e': 'èéêë',
            'i': 'ìíîï',
            'n': 'ñ',
            'o': 'òóôõö',
            'u': 'ùúûü',
            'y': 'ýÿ',
            's': 'š',
            'z': 'ž'
        }
    else:
        umlauts = {
            # 오류 우회
            '선': 'ퟴ',
            '휠': 'ퟵ',
            '젠': 'ퟶ',
            '츠': 'ퟷ',
            '유': 'ퟸ',
            '퀠': 'ퟹ',
            '술': 'ퟺ',
            '갠': 'ퟻ',
            ' ': '　',

            '년': [15],
            '일': [14],

            'A': 'ÀÁÂÃÄÅàáâãäå',
            'C': 'Çç',
            'E': 'ÈÉÊËèéêë',
            'I': 'ÌÍÎÏìíîï',
            'N': 'Ññ',
            'O': 'ÒÓÔÕÖòóôõö',
            'U': 'ÙÚÛÜùúûü',
            'Y': 'ÝŸýÿ',
            'S': 'Šš',
            'Z': 'Žž'
        }

    for c in range(0x21, 0x7F):
        if chr(c) in umlauts:
            umlauts[chr(c)] += chr(0xFEE0 + c)
        else:
            umlauts[chr(c)] = chr(0xFEE0 + c)

    char_id_re = re.compile('id=([0-9]*)')
    add_lines = []
    for line in lines:
        line = line.strip()
        m = char_id_re.search(line)
        if m:
            original_id = int(m.group(1))
            if chr(original_id) in umlauts:
                for char in umlauts[chr(original_id)]:
                    if type(char) == str:
                        char = ord(char)
                    add_lines.append(
                        line.replace('id=' + str(original_id), 'id=' + str(char))
                    )
    lines += add_lines

    bmfc_option = get_bmfc_option(name, ext)
    yoffset_modifier = bmfc_option.get('yOffset', 0)
    yoffset_re = re.compile('yoffset=([-0-9]*)')

    write_mode = False
    for line in lines:
        line = line.strip()
        if 'chars count=' in line:
            write_mode = True
            chars_count = int(line.split('=')[1])
            chars_count += len(add_lines)
            line = 'chars count=' + str(chars_count)

        m = yoffset_re.search(line)
        if m:
            original_offset = int(m.group(1))
            line = line.replace('yoffset=' + str(original_offset), 'yoffset=' + str(original_offset + yoffset_modifier))

        '''
        if ext == '._bmfc' and 'char id' in line:
            tokens = line.split(' ')
            char_id = int(tokens[1].split('=')[1])
            if char_id < 128:
                line = "char id=%d x=1  y=1  width=2    height=2    xoffset=1    yoffset=1    xadvance=1    page=0  chnl=15" % char_id
        '''
        f.write(line + '\n')

    f.close()

def check_int(s):
    if not s:
        return False
    if s[0] in ('-', '+'):
        return s[1:].isdigit()
    return s.isdigit()

def get_bmfc_option(name, ext):
    f = open('./original/bmfc/' + name + ext, 'r', encoding='utf_8')
    lines = f.readlines()

    bmfc_option = {}

    for line in lines:
        line = line.strip()
        if line and line[0] == '#':
            continue
        if '=' in line:
            elems = line.split('=')
            if len(elems) == 2:
                if check_int(elems[1]):
                    bmfc_option[elems[0]] = int(elems[1])
                else:
                    bmfc_option[elems[0]] = elems[1]
    return bmfc_option


def generate_dirs(position, width, height, unit):
    tmp_result = [position - unit, position - width * unit, position + unit, position + width * unit]
    result = []
    for value in tmp_result:
        if value >= 0 and value < width * height * unit:
            result.append(value)
    return result


def outglow_bmfont(name, ext, bmfc_option):
    print('Generating outglow.')
    f = open('./result/fonts/' + name + ext, 'rb')
    header = f.read(0x80)
    body = f.read()
    f.close()

    width = bmfc_option['outWidth']
    height = bmfc_option['outHeight']

    if len(body) != width * height * 4:
        print('The size of this DDS file is strange.')
        return

    generations = []
    points = set()

    for generation_number in range(0, 7):
        generations.append(set())
        current_generation = generations[-1]
        if generation_number == 0:
            for i in range(0, len(body), 4):
                r = body[i]

                if r == 255:  # Check not black position
                    dirs = generate_dirs(i, width, height, 4)
                    for p in dirs:
                        if body[p] != 255:
                            current_generation.add(i)
                            points.add(i)
                            break
        if generation_number != 0:
            for p in previous_generation:
                dirs = generate_dirs(p, width, height, 4)
                for i in dirs:
                    a = body[i]
                    if a == 255:
                        if i in points:
                            continue
                        current_generation.add(i)
                        points.add(i)
        previous_generation = current_generation

    result = bytearray(body)

    for i in range(0, len(result), 4):
        if result[i] == 255 and result[i + 3] != 0 and i not in points:
            result[i + 3] = 0
        if result[i] != 255:
            result[i + 3] = 0xBF
            if result[i] != 0:
                result[i] = result[i] // 2
                result[i + 1] = result[i + 1] // 2
                result[i + 2] = result[i + 2] // 2

    for generation_number, generation in enumerate(generations):
        alpha = [0xA9, 0x7A, 0x64, 0x4F, 0x3B, 0x28, 0x12][generation_number]
        for position in generation:
            result[position] = 0x91
            result[position + 1] = 0x91
            result[position + 2] = 0x91
            result[position + 3] = alpha

    f = open('./result/fonts/' + name + ext, 'wb')
    f.write(header)
    f.write(result)
    f.close()
    print('Finished.')

def generate_fonts():
    files = os.listdir("./original/bmfc")
    for file in files:
        file_name, file_ext = os.path.splitext(file)
        if file_ext == ".bmfc":
            generate_bmfont(file_name, file_ext, "ingame_source.txt");
        if file_ext == "._bmfc":
            generate_bmfont(file_name, file_ext, "worldmap_source.txt");
            bmfc_option = get_bmfc_option(file_name, file_ext)
            if bmfc_option['textureFormat'] != 'dds' or bmfc_option['textureCompression'] != 0:
                print('Auto glow is only possible for uncompressed DDS.')
                continue
            if bmfc_option['invR'] != 1 or bmfc_option['invG'] != 1 or bmfc_option['invB'] != 1:
                print('The Auto glow function requires that the invR, invG, and invB options are 1s.')
                continue

            fonts = os.listdir('./result/fonts')
            for font in fonts:
                font_name, font_ext = os.path.splitext(font)
                if font_name == file_name and font_ext == '.dds':
                    outglow_bmfont(font_name, font_ext, bmfc_option)

global_translated = {}

bypass_keywords = {
    ' ': '　',
    '선': 'ퟴ',
    '휠': 'ퟵ',
    '젠': 'ퟶ',
    '츠': 'ퟷ',
    '유': 'ퟸ',
    '퀠': 'ퟹ',
    '술': 'ퟺ',
    '갠': 'ퟻ',
}

hangul_re = re.compile('[가-힣]')

def bypass_error(src, is_map=False):
    res = src
    for key in bypass_keywords:
        res = res.replace(key, bypass_keywords[key])

    return res

def worksheet_to_yml(sheet_name, ws):
    header = {}
    translateds = {}
    origs = {}
    for idx, row in enumerate(ws.rows):
        if idx == 0:
            for jdx, col in enumerate(row):
                header[col.value] = jdx
        else:
            code = ws.cell(row=idx + 1, column=header['코드'] + 1).value
            orig = ws.cell(row=idx + 1, column=header['원문'] + 1).value
            text = ws.cell(row=idx + 1, column=header['역어'] + 1).value
            if code is None or orig is None or text is None:
                continue
            '''
            if orig in global_translated and text != global_translated[orig] and text == orig:
                print("%s> %s(%s) => %s" % (sheet_name, orig, text, global_translated[orig]))
            elif orig != text:
                global_translated[orig] = text
            '''
            extract_characters(None, text)
            text = bypass_error(text, True)
            translateds[code] = text
            origs[code] = orig

    f = open('./original/yml/' + sheet_name + '.yml', 'r', encoding='utf_8_sig')
    lines = f.readlines()
    f.close()

    os.makedirs('unified/yml', exist_ok=True)
    f = open('./unified/yml/' + sheet_name + '.yml', 'w', encoding='utf_8_sig')
    for line in lines:
        line = line.strip()
        tokens = line.split(' ', 1)
        if len(tokens) == 1:
            f.write(line + '\n')
        else:
            tokens[1] = tokens[1][1:-1]
            text = translateds.get(tokens[0], tokens[1])
            orig = origs.get(tokens[0], tokens[1])
            f.write(' %s "%s"\n' % (tokens[0], text))
    f.close()

def load_text_worksheet(text_translate_data, sheet_name, ws):
    header = {}
    for idx, row in enumerate(ws.rows):
        if idx == 0:
            for jdx, col in enumerate(row):
                header[col.value] = jdx
        else:
            category = ws.cell(row=idx + 1, column=header['분류'] + 1).value
            filename = ws.cell(row=idx + 1, column=header['파일명'] + 1).value
            orig = ws.cell(row=idx + 1, column=header['원문'] + 1).value
            text = ws.cell(row=idx + 1, column=header['역어'] + 1).value
            if category is None or filename is None or orig is None or text is None or orig == text:
                continue
            '''
            if orig in global_translated and text != global_translated[orig] and text == orig:
                print("%s> %s(%s) => %s" % (sheet_name, orig, text, global_translated[orig]))
            elif orig != text:
                global_translated[orig] = text
            '''
            extract_characters(None, text)
            text = bypass_error(text, True)
            if category not in text_translate_data:
                text_translate_data[category] = {}
            if filename not in text_translate_data[category]:
                text_translate_data[category][filename] = {}
            text_translate_data[category][filename][orig] = text

def unify_text_data(text_translate_data):
    filenames = set()
    for category in text_translate_data:
        if os.path.isdir('original/txt/%s' % category):
            for filename in text_translate_data[category]:
                if os.path.isfile('original/txt/%s/%s.txt' % (category, filename)):
                    filenames.add((category, filename))

    for (category, filename) in filenames:
        f = open('original/txt/%s/%s.txt' % (category, filename), 'r', encoding='ISO-8859-1')
        data = f.read()
        f.close()

        os.makedirs('unified/txt/%s' % (category), exist_ok=True)
        f = open('unified/txt/%s/%s.txt' % (category, filename), 'w', encoding='utf-8-sig')
        for key in text_translate_data[category][filename]:
            value = text_translate_data[category][filename][key]
            data = data.replace('"%s"' % key, '"%s"' % value)
        f.write(data)
        f.close()


def spreadsheet_to_data():
    print('read spreadsheets...')
    import openpyxl
    wb = openpyxl.load_workbook(filename='original.xlsx', data_only=True)
    sheets = wb.sheetnames
    text_translate_data = {}
    for sheet_name in sheets:
        if '_l_english' in sheet_name:
            print('Processing ' + sheet_name)
            ws = wb[sheet_name]
            worksheet_to_yml(sheet_name, ws)
        if '_txt' in sheet_name:
            print('Processing ' + sheet_name)
            ws = wb[sheet_name]
            load_text_worksheet(text_translate_data, sheet_name, ws)
    print('processing text datas...')
    unify_text_data(text_translate_data)
    print('done.')

def parse_podata(data):
    result = {}
    data = data.replace('"\n"', '')
    lines = data.split('\n')
    key = None
    msgctxt_re = re.compile('msgctxt "(.*)"')
    msgstr_re = re.compile('msgstr "(.*)"')

    for line in lines:
        line = line.strip()
        if 'msgctxt' in line:
            m = msgctxt_re.match(line)
            if m:
                key = m.group(1).replace('코', ':')
        if 'msgstr' in line:
            m = msgstr_re.match(line)
            if m and key is not None:
                value = m.group(1)
                result[key] = value
                key = None
    return result


def migrate_pofile(fn):
    f = open('./original/pofiles/%s.po' % fn, 'r', encoding='utf_8_sig')
    data = f.read()
    f.close()

    parsed_podata = parse_podata(data)

    f = open('./original/yml/%s.yml' % fn, 'r', encoding='utf_8_sig')
    lines = f.readlines()
    f.close()

    f = open('./original/yml/%s.yml' % fn, 'w', encoding='utf_8_sig')
    for line in lines:
        line = line.strip()
        tokens = line.split(' ', 1)
        if len(tokens) != 2:
            f.write(line + '\n')
        else:
            key, value = tokens
            if key in parsed_podata:
                f.write(' %s "%s"\n' % (key, parsed_podata[key]))
            else:
                f.write(' %s\n' % line)
    f.close()

def migrate_pofiles():
    print('migrate pofiles...')
    files = os.listdir('./original/pofiles/')
    for filename in files:
        fn, ext = os.path.splitext(filename)
        if ext == '.po':
            migrate_pofile(fn)

if __name__ == "__main__":
    # execute only if run as a script
    if len(sys.argv) == 1:
        is_unified = False
        if os.path.isdir('./original/pofiles/'):
            migrate_pofiles()
        if os.path.isfile('original.xlsx'):
            spreadsheet_to_data()
            is_unified = True

        escape_yml(is_unified)
        escape_txt(is_unified)
        generate_fonts()
    elif sys.argv[1] == '-u':
        is_unified = False
        if os.path.isdir('./original/pofiles/'):
            migrate_pofiles()
        if os.path.isfile('original.xlsx'):
            spreadsheet_to_data()
            is_unified = True
    elif sys.argv[1] == '-y':
        is_unified = False
        if os.path.isfile('original.xlsx'):
            spreadsheet_to_data()
            is_unified = True

        escape_yml(is_unified)
        escape_txt(is_unified)
    elif sys.argv[1] == '-f':
        generate_fonts()
    elif sys.argv[1] == '-n':
        pass
    else:
        print('Usage: (EU4Tools.py | EU4Tools.py -y | EU4Tools.py -f)')

# print(''.join(list(errornous_characters)))